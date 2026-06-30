#!/usr/bin/env -S uv run --quiet
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""xlsx-form — 既存 .xlsx フォームへの「最小改変」記入とハイライト検出。

哲学: フォーム xlsx は「図形・画像・共有文字列・印刷設定」を含む zip。
openpyxl の load→save は それらを落とし、Apple Numbers が開けなくなる
("The file format is invalid")。本ツールは **zip 内の該当セル XML だけ**を
書き換え、他パーツをバイト単位で温存する純粋変換 (filter)。

  highlights <in.xlsx>                 塗りつぶし(テーマ色含む)セルを列挙
  fill <in.xlsx> <spec.json> [out]     spec のセルを埋める。out 省略で stdout
       <in.xlsx> に "-" で stdin から読む。

spec.json 例 (シート名 → {セル参照: 値}):
  { "C　所属機関用１J": { "T17": 4, "G39": "〒116-0013 …", "AE32": 14 },
    "B　申請人用２Ｊ ": { "C12": "高木 俊輔" } }
値が int/float → 数値セル / str → inlineStr 文字列セル (電話番号等ハイフン入りは str に)。
"""
import sys, re, json, io, zipfile

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


# ---- 純粋ヘルパ (参照 <-> 列番号) ----
col_to_idx = lambda c: sum((ord(ch) - 64) * 26**i for i, ch in enumerate(reversed(c)))
parse_ref = lambda r: (re.match(r"([A-Z]+)(\d+)", r).group(1, 2))
xml_esc = lambda s: s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def cell_xml(ref, s_attr, value):
    s = f' s="{s_attr}"' if s_attr is not None else ""
    return (
        f'<c r="{ref}"{s} t="inlineStr"><is><t xml:space="preserve">{xml_esc(str(value))}</t></is></c>'
        if isinstance(value, str)
        else f'<c r="{ref}"{s}><v>{value}</v></c>'
    )


def read_bytes(path):
    return sys.stdin.buffer.read() if path == "-" else open(path, "rb").read()


def sheetname_to_part(zf):
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', rels))
    name_to_rid = re.findall(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', wb)
    return {name: "xl/" + rid_to_target[rid] for name, rid in name_to_rid if rid in rid_to_target}


# ---- fill: 該当セルだけ XML 置換 (存在しなければ行に挿入) ----
def edit_sheet_xml(xml, edits):
    for ref, value in edits.items():
        pat = re.compile(r'<c r="%s"((?:\s[^>]*?)?)(?:/>|>.*?</c>)' % re.escape(ref))
        m = pat.search(xml)
        if m:
            s = (re.search(r'\bs="(\d+)"', m.group(1)) or [None, None])[1]
            xml = xml[: m.start()] + cell_xml(ref, s, value) + xml[m.end() :]
            continue
        # 行に挿入 (列順を維持)。行が無ければ sheetData に行ごと挿入。
        col, row = parse_ref(ref)
        cidx = col_to_idx(col)
        rowpat = re.compile(r'(<row r="%s"(?:\s[^>]*)?>)(.*?)(</row>)' % row, re.S)
        rm = rowpat.search(xml)
        new_c = cell_xml(ref, None, value)
        if rm:
            cells = re.findall(r"<c [^>]*?(?:/>|>.*?</c>)", rm.group(2), re.S)
            pos = next((i for i, c in enumerate(cells)
                        if col_to_idx(parse_ref(re.search(r'r="([A-Z]+\d+)"', c).group(1))[0]) > cidx),
                       len(cells))
            cells.insert(pos, new_c)
            xml = xml[: rm.start()] + rm.group(1) + "".join(cells) + rm.group(3) + xml[rm.end() :]
        else:
            sdm = re.search(r"(<sheetData>)(.*?)(</sheetData>)", xml, re.S)
            rows = re.findall(r"<row [^>]*?(?:/>|>.*?</row>)", sdm.group(2), re.S)
            rpos = next((i for i, rr in enumerate(rows)
                         if int(re.search(r'r="(\d+)"', rr).group(1)) > int(row)), len(rows))
            rows.insert(rpos, f'<row r="{row}">{new_c}</row>')
            xml = xml[: sdm.start()] + sdm.group(1) + "".join(rows) + sdm.group(3) + xml[sdm.end() :]
    return xml


def fill(in_path, spec, out_path):
    raw = read_bytes(in_path)
    zin = zipfile.ZipFile(io.BytesIO(raw))
    name2part = sheetname_to_part(zin)
    # part(sheetN.xml) -> {ref: value}
    by_part = {}
    for sheet, edits in spec.items():
        part = sheet if sheet.startswith("xl/") else name2part.get(sheet)
        if not part:
            sys.exit(f"sheet not found: {sheet!r} (available: {list(name2part)})")
        by_part.setdefault(part, {}).update(edits)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():  # 元の順序・名前を維持
            data = zin.read(item.filename)
            if item.filename in by_part:
                data = edit_sheet_xml(data.decode("utf-8"), by_part[item.filename]).encode("utf-8")
            zout.writestr(item, data)
    out = buf.getvalue()
    (sys.stdout.buffer.write(out) if out_path in (None, "-") else open(out_path, "wb").write(out))
    if out_path not in (None, "-"):
        print(f"wrote {out_path} ({len(out)} bytes, {len(zin.infolist())} parts preserved)", file=sys.stderr)


# ---- highlights: 塗りつぶしセル列挙 (テーマ色解決) ----
def resolve_theme(path):
    import openpyxl  # noqa
    z = zipfile.ZipFile(path)
    t = z.read("xl/theme/theme1.xml").decode("utf-8")
    s = re.search(r"<a:clrScheme.*?</a:clrScheme>", t, re.S).group(0)
    els = re.findall(r"<a:(\w+)>(.*?)</a:\1>", s, re.S)
    d = {}
    for name, inner in els:
        mm = re.search(r'srgbClr val="([0-9A-Fa-f]{6})"', inner) or re.search(r'sysClr[^>]*lastClr="([0-9A-Fa-f]{6})"', inner)
        d[name] = mm.group(1) if mm else "?"
    # cell theme index: lt1/dk1 が clrScheme と入れ替わる
    order = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"]
    return {i: (order[i], d.get(order[i], "?")) for i in range(len(order))}


def highlights(path):
    import openpyxl
    theme = resolve_theme(path)
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        rows = []
        merged = list(ws.merged_cells.ranges)
        for row in ws.iter_rows():
            for c in row:
                f = c.fill
                if not (f and f.patternType):
                    continue
                fg = f.fgColor
                colr = (f"theme{fg.theme}->{theme.get(fg.theme,('?','?'))[0]}#{theme.get(fg.theme,('?','?'))[1]} tint{round(getattr(fg,'tint',0),2)}"
                        if getattr(fg, "type", None) == "theme"
                        else (fg.rgb if getattr(fg, "type", None) == "rgb" else str(fg.type)))
                if getattr(fg, "type", None) == "theme" and round(getattr(fg, "tint", 0), 3) != 0.0:
                    continue  # tint付き(薄い背景)は除外、ベタ塗りのみ
                label = next((ws.cell(row=c.row, column=cc).value for cc in range(1, c.column)
                              if ws.cell(row=c.row, column=cc).value not in (None, "")), None)
                mr = next((str(m) for m in merged if c.coordinate in m), None)
                rows.append((c.coordinate, colr, repr(c.value), mr, repr(label)))
        if rows:
            print(f"\n### {ws.title!r}")
            for r in rows:
                print(f"  {r[0]:6} fill={r[1]} val={r[2]} merged={r[3]} rowlabel={r[4]}")


def main():
    a = sys.argv[1:]
    if not a or a[0] in ("-h", "--help"):
        print(__doc__); return
    if a[0] == "highlights":
        highlights(a[1])
    elif a[0] == "fill":
        spec = json.load(open(a[2], encoding="utf-8"))
        fill(a[1], spec, a[3] if len(a) > 3 else None)
    else:
        sys.exit(f"unknown command: {a[0]} (use: highlights | fill)")


main() if __name__ == "__main__" else None
